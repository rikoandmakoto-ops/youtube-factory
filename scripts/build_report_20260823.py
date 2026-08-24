#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""YouTube Factory 日次分析レポート生成 2026-08-23（Phase 5）"""
import sqlite3
import re
import math
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

ROOT = Path(__file__).resolve().parent.parent
DB = ROOT / "data" / "analytics" / "analytics.db"
OUT = ROOT / "reports" / "youtube_analysis_20260823.xlsx"
SNAP = "2026-08-22"        # video_metrics 最新日
REACH_SNAP = "2026-08-20"  # video_reach_daily 最新日
ACTIVE = ["scp-lab", "daily-science", "pokemon-lab", "yokai-watch", "2ch-matome"]

FONT = "Arial"
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
SUB_FONT = Font(name=FONT, bold=True, size=11, color="1F3864")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="808080")
GOOD = PatternFill("solid", fgColor="E2EFDA")
BAD = PatternFill("solid", fgColor="FCE4E4")
WARN = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

con = sqlite3.connect(DB)
con.row_factory = sqlite3.Row


# ------------------------------------------------------------------ データ取得
def latest_videos(days=60):
    q = """with latest as (select video_id,max(date) d from video_metrics group by video_id)
    select m.* from video_metrics m join latest l
      on m.video_id=l.video_id and m.date=l.d
    where m.published_at>=date(?,'-%d day')""" % days
    return [dict(r) for r in con.execute(q, (SNAP,))]


VIDS = latest_videos()
REACH = defaultdict(lambda: [0, 0.0])
for vid, ch, imp, clk in con.execute(
    "select video_id,channel_id,sum(impressions),sum(clicks) from video_reach_daily group by video_id"
):
    REACH[vid] = [imp or 0, clk or 0.0]

NET30 = {
    r["channel_id"]: dict(r)
    for r in con.execute(
        """select channel_id,sum(subscribers_gained) g,sum(subscribers_lost) l,
           sum(subscribers_gained)-sum(subscribers_lost) net, sum(views) v
           from channel_metrics where date>=date('2026-08-19','-30 day') group by channel_id"""
    )
}

CH_REACH = {
    r["channel_id"]: dict(r)
    for r in con.execute(
        """select channel_id,sum(impressions) imp,sum(clicks) clk
           from video_reach_daily group by channel_id"""
    )
}


def ch_rows(ch):
    return [v for v in VIDS if v["channel_id"] == ch and (v["views"] or 0) > 0]


# ------------------------------------------------------------------ 共通ヘルパ
def header(ws, row, cols, widths=None):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill, cell.font, cell.border = H_FILL, H_FONT, THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


def put(ws, row, values, fmt=None, font=None, fills=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = font or BODY
        c.border = THIN
        if fmt and i - 1 < len(fmt) and fmt[i - 1]:
            c.number_format = fmt[i - 1]
        if fills and i - 1 < len(fills) and fills[i - 1]:
            c.fill = fills[i - 1]


wb = Workbook()

# ==================================================================== サマリー
ws = wb.active
ws.title = "サマリー"
ws.sheet_view.showGridLines = False
ws["A1"] = "YouTube Factory 日次分析レポート 2026-08-23"
ws["A1"].font = TITLE_FONT
ws["A2"] = (
    f"データ元: data/analytics/analytics.db（video_metrics 最新 {SNAP} / "
    f"video_reach_daily 最新 {REACH_SNAP}）　至上目標: チャンネル登録者数の増加"
)
ws["A2"].font = NOTE
ws["A3"] = (
    "※ 本日はじめて impressions / CTR が取得できたため、前日まで再生数のみで判定していた"
    "タイトル施策を CTR と登録者で再検証し、逆効果と判明したものを撤回した。"
)
ws["A3"].font = Font(name=FONT, size=10, bold=True, color="C00000")

ws["A5"] = "■ チャンネル別サマリー（直近60日公開分 / 純増は直近30日）"
ws["A5"].font = SUB_FONT
ws["A4"] = (
    "※ 本数・総再生・維持率・獲得登録者は「直近60日公開・再生数1以上」のコホート。"
    "インプレッションとCTRは reach レポートの全期間・全動画（母集団が異なるため直接の割り算はできない）。"
    "reach の取り込みは2日遅れのため、CTRは 2026-08-20 までの反映。"
)
ws["A4"].font = NOTE

cols = ["チャンネル", "本数", "総再生", "平均再生", "維持率", "インプレッション\n(全期間)", "CTR\n(全期間)",
        "獲得登録者", "登録者/本", "登録者/1000再生", "30日純増"]
header(ws, 6, cols, [16, 7, 10, 10, 9, 15, 8, 11, 11, 15, 10])

r = 7
first_data = r
for ch in ACTIVE:
    rows = ch_rows(ch)
    n = len(rows)
    tv = sum(v["views"] or 0 for v in rows)
    ret = sum(v["avg_view_percentage"] or 0 for v in rows) / max(n, 1) / 100
    subs = sum(v["subscribers_gained"] or 0 for v in rows)
    imp = (CH_REACH.get(ch) or {}).get("imp") or 0
    clk = (CH_REACH.get(ch) or {}).get("clk") or 0
    net = (NET30.get(ch) or {}).get("net") or 0
    ws.cell(row=r, column=1, value=ch)
    ws.cell(row=r, column=2, value=n)
    ws.cell(row=r, column=3, value=tv)
    ws.cell(row=r, column=4, value=f"=IF(B{r}=0,0,C{r}/B{r})")
    ws.cell(row=r, column=5, value=ret)
    ws.cell(row=r, column=6, value=imp)
    ws.cell(row=r, column=7, value=(clk / imp) if imp else 0)
    ws.cell(row=r, column=8, value=subs)
    ws.cell(row=r, column=9, value=f"=IF(B{r}=0,0,H{r}/B{r})")
    ws.cell(row=r, column=10, value=f"=IF(C{r}=0,0,H{r}/C{r}*1000)")
    ws.cell(row=r, column=11, value=net)
    put(ws, r, [None] * 11,
        fmt=[None, "#,##0", "#,##0", "#,##0", "0.0%", "#,##0", "0.00%", "#,##0", "0.000", "0.000", "+#,##0;-#,##0"])
    for cc in range(1, 12):
        ws.cell(row=r, column=cc).font = BODY
        ws.cell(row=r, column=cc).border = THIN
    ws.cell(row=r, column=1).font = BOLD
    r += 1
last_data = r - 1

ws.cell(row=r, column=1, value="合計/加重平均").font = BOLD
for col, f in [
    (2, f"=SUM(B{first_data}:B{last_data})"),
    (3, f"=SUM(C{first_data}:C{last_data})"),
    (4, f"=IF(B{r}=0,0,C{r}/B{r})"),
    (5, f"=SUMPRODUCT(B{first_data}:B{last_data},E{first_data}:E{last_data})/B{r}"),
    (6, f"=SUM(F{first_data}:F{last_data})"),
    (7, f"=SUMPRODUCT(F{first_data}:F{last_data},G{first_data}:G{last_data})/F{r}"),
    (8, f"=SUM(H{first_data}:H{last_data})"),
    (9, f"=IF(B{r}=0,0,H{r}/B{r})"),
    (10, f"=IF(C{r}=0,0,H{r}/C{r}*1000)"),
    (11, f"=SUM(K{first_data}:K{last_data})"),
]:
    ws.cell(row=r, column=col, value=f)
put(ws, r, [None] * 11,
    fmt=[None, "#,##0", "#,##0", "#,##0", "0.0%", "#,##0", "0.00%", "#,##0", "0.000", "0.000", "+#,##0;-#,##0"],
    font=BOLD)
ws.cell(row=r, column=1, value="合計/加重平均").font = BOLD
total_row = r

ws.cell(row=first_data, column=10).comment = Comment(
    "scp-lab の登録者/1000再生は他チャンネルの約2.6倍。"
    "リーチ(インプレッション)も最大だが CTR は最下位で、クリック率の改善が"
    "全社の登録者数に最も効く構造になっている。", "指揮者", height=110, width=320)

r += 2
ws.cell(row=r, column=1, value="■ 本日の重要な発見").font = SUB_FONT
r += 1
findings = [
    ("1", "CTR/インプレッションの取得が復旧（前日の最重要未解決事項）",
     "video_reach_daily に 57,887 インプレッション・909 クリックが蓄積され、はじめて CTR ベースの評価が可能になった。"
     "これにより、再生数だけで判定していた前日までのタイトル施策を検証できるようになった。"),
    ("2", "前日のタイトル施策5件のうち4件が、登録者ベースでは支持されず撤回",
     "yokai-watch の「怖」優先（登録者0.75倍）、yokai-watch の「なぜ」禁止（1.12倍でほぼ中立）、"
     "daily-science の「なぜ」抑制（1.58倍で有効）、daily-science の「99%が知らない」禁止（1.57倍で有効）。"
     "scp-lab のダッシュ「必ず維持」も、同一コホートでは有意差がなく根拠がなかった。"
     "5件すべてが再生数だけを見て決めたもの。詳細は「改善アクション」シート。"),
    ("3", "【根本原因】設定の「書き込み先が違う」問題が2箇所あった",
     "(a) テーマ補充プロンプト(generator.suggest_themes)は voice_style.style_rules を一切参照せず "
     "theme_priority.title_style / good_examples のみを使う。そのため前日のタイトル施策は生成に無効だった。"
     "(b) テーマキューが2系統あり、定時投稿の autopilot は <channel>.json の autopilot.theme_queue を読む。"
     "data/channels/<id>/theme_queue.json は手動の /factory/run 専用。"
     "本日はまず(b)の使われない方だけを直しており、自己検証で気づいて両方を是正した。"),
    ("4", "維持率の急落は見かけ。公開後経過日数を揃えると崩壊していない",
     "週次集計では全5chが W33 で維持率20-36%まで落ちたが、これは公開直後の動画がデータ未成熟なため。"
     "公開後3日で揃えると 41-55% で、実際の急落はない。ただし pokemon-lab と yokai-watch は "
     "W30→W33 で 78.7%→55.0% / 63.3%→50.2% と緩やかな低下傾向がある。"),
    ("5", "scp-lab がボトルネックかつ最大の伸びしろ",
     "登録者の54%を1chで稼ぐ一方、CTR 1.52%（最下位）・30%地点の残存0.64（最下位）。"
     "リーチは全ch中最大の43,997インプレッションを持つため、CTR と維持率の改善が最も費用対効果が高い。"),
    ("6", "短い尺が登録者に効くが、効果量は当初見えたより小さい",
     "単純な三分位比較では短尺が2-3倍に見えたが、これは推定尺(視聴秒数÷維持率)と維持率の自己相関による過大評価。"
     "維持率45-70%に絞って再検証すると 1.33倍（95%CI 0.88-2.01）で有意ではない。過度な短尺化は行わない。"),
    ("7", "【自己検証で判明】本日の結論そのものに4件の誤りがあり、公開前に訂正した",
     "(a) CTRの集計母集団が登録者指標と食い違っていた（CTRは全期間、登録者は60日コホート）。"
     "揃え直した結果、本日の目玉だった『ダッシュ—は0.72倍で有害』は 0.97倍（95%CI 0.81-1.16）となり撤回。"
     "(b) scp-lab「怖/恐」のCTR 1.26倍も有意でなく、必須から推奨へ格下げ。"
     "(c) テーマキューに実在しない数字（「9つ」「26部隊」「4名」）を付与していたため全削除。"
     "(d) 既存PDCAメモリと突合し、daily-science の「99%が知らない」禁止が登録者ベースでは誤りと判明、解除。"),
    ("8", "同じ誤りを繰り返した構造的な原因は「再生数と登録者の食い違い」を都度確かめていないこと",
     "前日・本日ともに、単一指標だけを見て方針を決めた結果が翌日に覆っている。"
     "再生数・CTR・登録者は頻繁に符号が食い違う（例: yokai-watch の伝承回は再生+21%だが登録-59%）。"
     "今後は必ず (1)集計コホートを揃える (2)再生・CTR・1本あたり登録者の3指標を併記する "
     "(3)95%信頼区間が1をまたぐ場合は『必須』にしない、の3点を満たしてから反映する。"),
]
for no, t, d in findings:
    ws.cell(row=r, column=1, value=no).font = BOLD
    ws.cell(row=r, column=2, value=t).font = BOLD
    r += 1
    ws.cell(row=r, column=2, value=d).font = BODY
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
    ws.row_dimensions[r].height = 42
    r += 1

# ============================================================ チャンネル別詳細
ws2 = wb.create_sheet("チャンネル別詳細")
ws2.sheet_view.showGridLines = False
ws2["A1"] = "チャンネル別詳細"
ws2["A1"].font = TITLE_FONT
row = 3

CURVE = defaultdict(lambda: defaultdict(list))
import json as _json
for ch, cv in con.execute("select channel_id,curve from retention_curve"):
    try:
        pts = _json.loads(cv)
    except Exception:
        continue
    for p in pts:
        rt, y = p.get("ratio"), p.get("audience_watch_ratio")
        if rt is None or y is None:
            continue
        CURVE[ch][round(float(rt) * 10) / 10].append(float(y))

COMPETITOR = defaultdict(list)
mx = con.execute("select max(analysis_date) from competitor_analyses").fetchone()[0]
for r_ in con.execute(
    "select channel_id,competitor_title,subscriber_count,avg_views,posting_frequency_per_week "
    "from competitor_analyses where analysis_date=? order by avg_views desc", (mx,)
):
    COMPETITOR[r_["channel_id"]].append(dict(r_))

for ch in ACTIVE:
    rows = ch_rows(ch)
    n = len(rows)
    ws2.cell(row=row, column=1, value=f"● {ch}").font = SUB_FONT
    row += 1
    tv = sum(v["views"] or 0 for v in rows)
    subs = sum(v["subscribers_gained"] or 0 for v in rows)
    imp = (CH_REACH.get(ch) or {}).get("imp") or 0
    clk = (CH_REACH.get(ch) or {}).get("clk") or 0
    net = NET30.get(ch) or {}
    stats = [
        ("公開本数(60日)", n, "#,##0"),
        ("総再生", tv, "#,##0"),
        ("平均再生/本", tv / max(n, 1), "#,##0"),
        ("平均維持率", sum(v["avg_view_percentage"] or 0 for v in rows) / max(n, 1) / 100, "0.0%"),
        ("平均視聴秒数", sum(v["avg_view_duration"] or 0 for v in rows) / max(n, 1), "0.0"),
        ("インプレッション", imp, "#,##0"),
        ("CTR", (clk / imp) if imp else 0, "0.00%"),
        ("獲得登録者(60日)", subs, "#,##0"),
        ("登録者/1000再生", (subs * 1000 / tv) if tv else 0, "0.000"),
        ("30日純増", net.get("net", 0), "+#,##0;-#,##0"),
        ("高評価率", sum((v["likes"] or 0) for v in rows) / max(tv, 1), "0.00%"),
        ("コメント総数", sum(v["comments"] or 0 for v in rows), "#,##0"),
    ]
    header(ws2, row, [s[0] for s in stats], [16] * len(stats))
    row += 1
    put(ws2, row, [s[1] for s in stats], fmt=[s[2] for s in stats])
    row += 2

    ws2.cell(row=row, column=1, value="視聴維持カーブ（audience_watch_ratio・1.00未満=平均以下）").font = BOLD
    row += 1
    pos = [i / 10 for i in range(0, 11)]
    header(ws2, row, ["再生位置"] + [f"{int(p*100)}%" for p in pos], [16] + [8] * 11)
    row += 1
    vals = ["平均残存"] + [
        (sum(CURVE[ch][p]) / len(CURVE[ch][p])) if CURVE[ch].get(p) else None for p in pos
    ]
    fills = [None] + [
        (BAD if (v is not None and v < 0.7) else (WARN if (v is not None and v < 1.0) else GOOD))
        for v in vals[1:]
    ]
    put(ws2, row, vals, fmt=[None] + ["0.00"] * 11, fills=fills)
    row += 2

    comps = COMPETITOR.get(ch) or []
    if comps:
        ws2.cell(row=row, column=1, value=f"競合比較（{mx} 時点・上位5）").font = BOLD
        row += 1
        header(ws2, row, ["競合チャンネル", "登録者数", "平均再生", "週間投稿本数", "自ch比(平均再生)"], [34, 12, 12, 14, 16])
        row += 1
        own = tv / max(n, 1)
        for cp in comps[:5]:
            put(ws2, row, [cp["competitor_title"], cp["subscriber_count"], cp["avg_views"],
                           cp["posting_frequency_per_week"],
                           (cp["avg_views"] / own) if own else 0],
                fmt=[None, "#,##0", "#,##0", "0.0", "0.0\"倍\""])
            row += 1
    row += 2

# ====================================================== 動画別パフォーマンス
ws3 = wb.create_sheet("動画別パフォーマンス")
ws3.sheet_view.showGridLines = False
ws3["A1"] = "動画別パフォーマンス（直近60日公開・インプレッション80件以上を対象にCTRを算出）"
ws3["A1"].font = TITLE_FONT
header(ws3, 3, ["チャンネル", "公開日", "タイトル", "再生数", "維持率", "視聴秒数",
                "インプレッション", "クリック", "CTR", "登録者", "登録者/1000再生", "ダッシュ—", "怖/恐", "秘密/隠"],
       [14, 11, 62, 9, 8, 9, 14, 9, 8, 8, 15, 9, 8, 8])
row = 4
allv = sorted(VIDS, key=lambda v: (v["channel_id"], -(v["subscribers_gained"] or 0), -(v["views"] or 0)))
for v in allv:
    if (v["views"] or 0) <= 0:
        continue
    imp, clk = REACH.get(v["video_id"], [0, 0])
    t = v["title"] or ""
    ctr = (clk / imp) if imp >= 80 else None
    put(ws3, row, [
        v["channel_id"], (v["published_at"] or "")[:10], t, v["views"],
        (v["avg_view_percentage"] or 0) / 100, v["avg_view_duration"],
        imp or None, (clk or None) if imp >= 80 else None, ctr,
        v["subscribers_gained"],
        (v["subscribers_gained"] or 0) * 1000 / (v["views"] or 1),
        "○" if "—" in t else "", "○" if re.search(r"[怖恐]", t) else "",
        "○" if re.search(r"秘密|隠", t) else "",
    ], fmt=[None, None, None, "#,##0", "0.0%", "0.0", "#,##0", "#,##0", "0.00%", "#,##0", "0.000", None, None, None])
    if ctr is not None:
        ws3.cell(row=row, column=9).fill = GOOD if ctr >= 0.025 else (BAD if ctr < 0.01 else WARN)
    row += 1
ws3.freeze_panes = "A4"
ws3.auto_filter.ref = f"A3:N{row-1}"

# ================================================================ 改善アクション
ws4 = wb.create_sheet("改善アクション")
ws4.sheet_view.showGridLines = False
ws4["A1"] = "改善アクション 2026-08-23（すべて実データに基づき、本日コンフィグへ反映済み）"
ws4["A1"].font = TITLE_FONT
header(ws4, 3, ["#", "対象", "変更内容", "根拠（実測値）", "種別", "反映先"],
       [4, 14, 44, 52, 12, 40])
row = 4
ACTIONS = [
    ("1", "scp-lab", "全角ダッシュ「—」を中立扱いに変更（前日の『必ず維持する』と、本日一度出した『使用禁止』の両方を撤回）",
     "全期間集計では CTR 0.72倍（95%CI 0.62-0.85）に見えたが、登録者指標と同じ直近60日コホートに揃えると "
     "CTR 0.97倍（95%CI 0.81-1.16）、1本あたり登録者 0.74倍（95%CI 0.47-1.18）でいずれも有意差なし。"
     "差は61日以上前の動画による期間交絡だった。夜間PDCAの独立分析（登録/1k 11.75 vs 11.38）とも一致する。",
     "撤回", "theme_priority.title_style / good_examples / voice_style.style_rules"),
    ("2", "scp-lab", "タイトルに「怖」「恐」を入れることを推奨（当初は必須化したが検証を受け推奨に格下げ）",
     "1本あたり登録者 0.923 vs 0.612 = 1.51倍（60日コホート）。"
     "※CTR単体では1.24倍（95%CI 0.99-1.55）で有意でなく、60日コホートに揃えると0.93倍と符号が反転する。"
     "根拠は最終成果である登録者指標のみとし、必須化はしない。翌週データで再検証する。",
     "新規(弱)", "theme_priority.title_style / good_examples"),
    ("3", "scp-lab", "タイトルの「本当」「実は」系を回避",
     "CTR 1.30% vs 1.61% = 0.81倍（95%CI 0.68-0.96）、1本あたり登録者 0.84倍。"
     "※測定述語は「本当」または「実は」を含むこと。60日コホートでは0.83倍（95%CI 0.69-1.01）で非有意。",
     "新規", "theme_priority.title_style"),
    ("4", "yokai-watch", "「怖」優先指定を撤回（前日追加分）",
     "CTR 0.89倍・1本あたり登録者 0.375 vs 0.500 = 0.75倍。再生数(1.24倍)は伸びるが登録に繋がらない。",
     "撤回", "theme_priority.title_style / voice_style.style_rules"),
    ("5", "yokai-watch", "「なぜ」始まりの使用禁止を解除（前日追加分）",
     "CTR 0.94倍・1本あたり登録者1.12倍でほぼ中立。前日は再生数0.80倍のみで禁止していた。",
     "撤回", "theme_priority.title_style / voice_style.style_rules"),
    ("6", "yokai-watch", "「秘密」「隠」「本当」「実は」のいずれかを必須化",
     "「秘密/隠」CTR 2.67% vs 1.34% = 1.99倍（95%CI 1.32-3.02）。「本当/実は」1本あたり登録者 1.00 vs 0.28 = 3.60倍（95%CI 1.10-11.80）。"
     "※後者はA群n=6・実登録6人と少数。測定述語は「本当」であって「本当は」に限定した効果ではない。",
     "新規", "theme_priority.title_style / good_examples"),
    ("7", "yokai-watch", "台本の尺を 320-390字 → 270-310字 に短縮（前日の延長を部分撤回）",
     "維持率45-70%に揃えた比較で、短尺側は1本あたり登録者0.875、長尺側0.375。前日は再生数(長尺1629 vs 短尺1160)を見て延長したが、登録者では逆だった。",
     "撤回", "short_format.total_chars_min / max / line_chars"),
    ("8", "daily-science", "「なぜ」始まりの使用抑制を解除",
     "1本あたり登録者 0.400 vs 0.253 = 1.58倍。前日は再生数0.95倍で『効果なし』と判定していた。"
     "※同コホートのCTRは 0.79% vs 1.48% = 0.53倍と負で、CTRと登録者で符号が食い違う。"
     "登録者/本は再生を経た最終成果のため後者を採用したが、サムネ側でクリック率を補う必要がある。",
     "撤回", "theme_priority.title_style / voice_style.style_rules"),
    ("9", "daily-science", "疑問符「？」を含む形を必須化、「本当」「実は」を禁止",
     "「？」1本あたり登録者 0.324 vs 0.150 = 2.16倍（95%CI 0.66-7.03、非該当群がn=20と少ない）。"
     "「本当/実は」0.241 vs 0.344 = 0.70倍。",
     "新規", "theme_priority.title_style / good_examples"),
    ("10", "pokemon-lab", "「秘密」「隠」型を毎バッチ最低1本必須化",
     "1本あたり登録者 1.00 vs 0.32 = 3.17倍（95%CI 1.02-9.82）。当チャンネルで唯一有意だが下限がぎりぎり1を超える弱い証拠。"
     "A群n=6・実登録6人のうち3本は登録0で、寄与が1本に集中している。翌週データで再検証が必要。",
     "新規(弱)", "theme_priority.title_style / good_examples / voice_style.style_rules"),
    ("11", "pokemon-lab", "『どっちが勝つ』対決型を1バッチ1本までに制限、締めに登録動機を明示",
     "対決型の1本あたり登録者は0.63倍。個別動画のCTRは3.82%/2.87%と高いが、群として比較すると "
     "2.12% vs 1.57% = 1.35倍（95%CI 0.69-2.64）で有意ではない。クリックは取れるが登録に繋がっていない。",
     "新規", "theme_priority.title_style / voice_style.style_rules"),
    ("12", "全チャンネル", "【根本原因対応】タイトル規則を theme_priority 側へ移設",
     "テーマ補充プロンプトは voice_style.style_rules を参照しない実装（generator.py:2755-2972）。前日の施策が生成に反映されず、scp-lab のキュー23件全てがダッシュ形式のままだった。",
     "根本対応", "theme_priority.title_style / good_examples（4ch）"),
    ("13", "scp-lab / yokai-watch / pokemon-lab", "既存テーマキュー40件のタイトルを新規則に沿って全件書き直し",
     "初回の書き直しでは (a) 実在しない数字の付与（「9つ」「26部隊」「4名」）、(b) 総称のみで固有名のないタイトル6件、"
     "(c) good_examples との重複4件、(d) 同語反復「機動部隊MTF」、(e) 伏字の不自然な活用「███手順」があり、"
     "独立検証を経て全て修正した。最終状態でルール適合・捏造数字なし・固有名あり・重複なしを機械検証済み。",
     "新規", "data/channels/*/theme_queue.json"),
    ("14", "yokai-watch", "テーマキューを 6件 → 10件（target_size）へ補充",
     "min_threshold=5 に接近していたため。追加4件はすべて新タイトル規則に適合。created_at も実日付に修正済み。",
     "新規", "data/channels/yokai-watch/theme_queue.json"),
    ("15", "generator.py", "テーマ補充プロンプトの例示からダッシュ「—」を除去",
     "プロンプト内の続編タイトル例（generator.py:2871）がダッシュ形式を例示しており、規則と矛盾していた。",
     "根本対応", "backend/pipeline/auto_scenario/generator.py"),
    ("16", "全チャンネル", "viral_hooks の破損（文字列が1文字ずつlistに分解）を修復",
     "scp-lab 90件・yokai-watch 69件・daily-science 159件・pokemon-lab 70件がすべて1文字要素。プロンプトを汚染していた。",
     "不具合修正", "theme_priority.viral_hooks"),
    ("17", "clip-lab", "autopilot を停止（enabled: true → false）",
     "運用方針で凍結中にもかかわらず平日2枠＋休日2枠がスケジュールされ続けていた。theme_seeds 0件・theme_queue 未作成・video_metrics 実績0件で空回りしていた。",
     "不具合修正", "autopilot.enabled"),
    ("18", "scp-lab", "台本が short_format の上限を超過している問題への対処規則を追加",
     "目標22秒に対し実測の推定尺は26.0-35.8秒。超過時は情報を足さず行を削って収める旨を明記。",
     "新規", "voice_style.style_rules"),
    ("19", "scp-lab / daily-science", "タイトルへの数字の付与を「実在する場合のみ」に変更",
     "当初は『数字を必ず1つ以上入れる』としたが、原典に数字がないテーマで数字が創作される事故が実際に発生した。"
     "報告書としての信頼性を損なうため、出典が確認できない数字の創作を明示的に禁止した。",
     "不具合修正", "theme_priority.title_style"),
    ("20", "daily-science", "viral_hooks を原文のまま維持（一度除去したが復元）",
     "当初は title_style の禁止と矛盾するとして「99%が知らない」を除去したが、"
     "(a) 除去処理の正規表現が viral_hooks 文字列全体を消してしまう不具合があり、"
     "(b) そもそも禁止自体が誤りだったため（アクション#22）、バックアップから全文を復元した。",
     "不具合修正", "theme_priority.viral_hooks"),
    ("21", "yokai-watch", "total_chars_max を 330 → 310 に修正",
     "line_chars（1-7行27〜37字、8行目+15字）から到達可能な上限は311字で、330字は算術的に到達不能だった。",
     "不具合修正", "short_format.total_chars_max"),
    ("22", "daily-science", "「99%が知らない」の禁止を解除（前日からの継続禁止を撤回）",
     "1本あたり登録者 0.417 vs 0.265（1.57倍）、登録/1000再生 0.47 vs 0.29、CTR 1.53% vs 1.28% と3指標すべてで正。"
     "前日は再生数0.94倍のみを理由に禁止していたが、至上目標は登録者であるため誤り。"
     "ただし95%CI 0.76-3.26 と有意ではないため必須化はせず、多用も避ける。",
     "撤回", "theme_priority.title_style / voice_style.style_rules"),
    ("24", "scp-lab / yokai-watch / pokemon-lab", "【本日最大の修正】autopilot が実際に消費するキューを是正",
     "テーマキューが2系統あり、定時投稿の autopilot は <channel>.json の autopilot.theme_queue を読む"
     "（api_channel_autopilot.py の _pop_or_refill_theme）。data/channels/<id>/theme_queue.json は手動実行の "
     "/factory/run 専用だった。反映前の autopilot.theme_queue は scp-lab 7件すべて「怖/恐」なし、"
     "yokai-watch 10件すべてダッシュ使用かつ全件が伝承のみ（作品ネタゼロ＝本日の最有力知見と正面から矛盾）、"
     "pokemon-lab 12件中5件が対決型。精査済みキューで置き換え、両系統を一致させた。",
     "根本対応", "autopilot.theme_queue（6ファイル）"),
    ("23", "yokai-watch", "「妖怪ウォッチ作品ネタ」を題材選定の最優先条件に格上げ",
     "作品名・キャラ名・ゲーム内仕様を含む10本は1本あたり登録者 0.700 vs 伝承のみ0.286（2.45倍）、"
     "登録/1000再生 2.9倍、CTR 1.96% vs 1.47%。伝承のみの回は平均再生では上回る（1543 vs 1281）が登録に繋がらない。"
     "秘密/隠よりも効果量が大きく、夜間PDCAの独立分析とも一致する。",
     "新規", "theme_priority.title_style / voice_style.style_rules"),
]
for a in ACTIONS:
    put(ws4, row, list(a))
    ws4.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
    _f = {"撤回": BAD, "新規": GOOD, "新規(弱)": WARN,
          "根本対応": WARN, "不具合修正": WARN}.get(a[4])
    if _f is not None:
        ws4.cell(row=row, column=5).fill = _f
    ws4.row_dimensions[row].height = 46
    row += 1

row += 1
ws4.cell(row=row, column=1, value="■ 未解決事項").font = SUB_FONT
row += 1
header(ws4, row, ["#", "内容", "状況", "次アクション"], [4, 44, 20, 60])
row += 1
OPEN = [
    ("1", "company-facts の実績データが0件のまま",
     "前日から継続", "autopilot は有効でテーマキュー9件もあるが video_metrics / channel_metrics ともに0行。投稿が実際にYouTubeへ到達しているかバックエンドログで確認が必要。"),
    ("2", "2ch-matome が最下位で改善の打ち手が立たない",
     "データ不足", "60日で公開23本（うち再生数1以上は17本）・獲得登録者3人。インプレッションも551件のみでCTR分析の検出力がない。データが貯まるまで施策変更を保留する（前日の1本/日への減枠は維持）。"),
    ("3", "video_metrics の impressions と video_reach_daily の値が一致しない",
     "新規", "scp-lab は video_metrics 29,585 に対し video_reach_daily 43,997。本レポートは母数の大きい video_reach_daily を採用したが、取り込みロジックの差分を確認すべき。"),
    ("4", "reach レポートの取り込みが2-3日遅れる",
     "新規", "video_reach_daily の最新が 2026-08-20 で video_metrics（08-22）より2日古い。当日投稿分のCTRは翌々日まで評価できない前提で運用する。"),
    ("5", "pokemon-lab と yokai-watch の維持率が4週連続で低下",
     "新規", "公開後3日で揃えた比較で pokemon-lab 78.7%→55.0%、yokai-watch 63.3%→50.2%。初期ブーストの減衰か品質低下かの切り分けが必要。翌日以降も継続監視。"),
    ("6", "指揮者からバックエンド(localhost:8000)へ到達できない",
     "前日から継続", "指揮者はサンドボックスLinux上で動作するため Mac の localhost に到達不可。コンフィグ更新が実質の制作指示となる（APScheduler が変更を自動で拾う）。即時実行は scripts/trigger_autopilot_20260823.sh をMacで実行。"),
    ("7", "本日のタイトル規則の効果は未検証",
     "新規", "reach 取り込みが2日遅れるため、本日の変更のCTR効果は 08-26 頃まで判定できない。それまで追加のタイトル変更は行わず、効果測定を優先する。"),
    ("8", "本日の施策のうち2件は統計的根拠が弱い",
     "新規", "scp-lab「怖/恐」（CTR非有意・60日コホートで符号反転）と pokemon-lab「秘密/隠」（A群n=6・寄与が1本に集中、CI下限1.02）は、いずれも弱い証拠に基づく。翌週データで再検証し、支持されなければ撤回する。"),
    ("9", "分析の母集団を統一する仕組みがない",
     "新規", "CTRは reach 全期間、登録者は60日コホートで集計しており、同一の主張内で母集団が混在しやすい。次回以降は集計コホートを定義した共通クエリを scripts/ に用意して揃える。"),
    ("10", "テーマキューが2系統あり、恒常的にドリフトする",
     "新規", "autopilot.theme_queue（定時投稿が消費）と data/channels/<id>/theme_queue.json（/factory/run が消費）が独立に存在し、補充ロジックも別（前者は _pop_or_refill_theme、後者は theme_queue.replenish）。本日3chは手動で一致させたが daily-science は中身が異なるまま。どちらか一方に寄せるか、同期処理をバックエンドに入れるべき。"),
    ("11", "pokemon-lab の設定が channels と orchestrator で食い違っている",
     "新規", "auto_optimize_schedule が channels=true / orchestrator=false、schedule.days_of_week も channels=[水木土] / orchestrator=[毎日]。前日に全5chを false 固定したはずだが channels 側が true に戻っている。夜間の別処理が書き戻している可能性があり、書き込み主体の特定が必要。"),
]
for o in OPEN:
    put(ws4, row, list(o))
    ws4.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[row].height = 46
    row += 1

# ==================================================================== トレンド
ws5 = wb.create_sheet("トレンド")
ws5.sheet_view.showGridLines = False
ws5["A1"] = "トレンド"
ws5["A1"].font = TITLE_FONT

ws5["A3"] = "■ 公開週別の推移（公開後3日時点で揃えた維持率。週次の生値は新着動画のデータ未成熟で過小に出るため）"
ws5["A3"].font = SUB_FONT
header(ws5, 4, ["チャンネル", "公開週", "本数", "平均再生(最新値)", "維持率(公開3日時点)", "獲得登録者"],
       [16, 12, 8, 18, 20, 12])
row = 5
q = """select channel_id, strftime('%Y-W%W',published_at) wk, count(*) n,
 round(avg(avg_view_percentage),1) ret from video_metrics
 where published_at>=date(?,'-42 day') and avg_view_percentage>0
 and cast(julianday(date)-julianday(substr(published_at,1,10)) as int)=3
 group by channel_id,wk"""
ret3 = {(r_["channel_id"], r_["wk"]): (r_["n"], r_["ret"]) for r_ in con.execute(q, (SNAP,))}
wk_agg = defaultdict(lambda: [0, 0, 0])
for v in VIDS:
    if not v["published_at"]:
        continue
    import datetime as _dt
    d = _dt.date.fromisoformat(v["published_at"][:10])
    wk = f"{d.year}-W{d.strftime('%W')}"
    a = wk_agg[(v["channel_id"], wk)]
    a[0] += 1
    a[1] += v["views"] or 0
    a[2] += v["subscribers_gained"] or 0
for (ch, wk), (n, tv, sg) in sorted(wk_agg.items()):
    if ch not in ACTIVE:
        continue
    r3 = ret3.get((ch, wk))
    put(ws5, row, [ch, wk, n, tv / max(n, 1),
                   (r3[1] / 100) if r3 else None, sg],
        fmt=[None, None, "#,##0", "#,##0", "0.0%", "#,##0"])
    row += 1

row += 2
ws5.cell(row=row, column=1, value="■ 未処理のトレンド検出（status=detected・スコア上位）").font = SUB_FONT
row += 1
header(ws5, row, ["チャンネル", "キーワード", "スコア", "提案タイトル", "検出日"], [16, 26, 9, 58, 20])
row += 1
for r_ in con.execute(
    """select channel_id,keyword,combined_score,suggested_title,detected_at
       from trend_detections where status='detected' and channel_id in
       ('scp-lab','daily-science','pokemon-lab','yokai-watch','2ch-matome')
       order by combined_score desc, detected_at desc limit 30"""
):
    put(ws5, row, [r_["channel_id"], r_["keyword"], r_["combined_score"],
                   r_["suggested_title"], str(r_["detected_at"] or "")[:19]],
        fmt=[None, None, "0.00", None, None])
    row += 1

row += 2
ws5.cell(row=row, column=1, value="■ トレンド検出のステータス別件数").font = SUB_FONT
row += 1
header(ws5, row, ["チャンネル", "detected(未処理)", "queued(投入済)", "合計"], [16, 16, 16, 10])
row += 1
st = defaultdict(lambda: defaultdict(int))
for ch, s, n in con.execute("select channel_id,status,count(*) from trend_detections group by channel_id,status"):
    st[ch][s] = n
for ch in ACTIVE:
    d_, q_ = st[ch].get("detected", 0), st[ch].get("queued", 0)
    ws5.cell(row=row, column=1, value=ch)
    ws5.cell(row=row, column=2, value=d_)
    ws5.cell(row=row, column=3, value=q_)
    ws5.cell(row=row, column=4, value=f"=B{row}+C{row}")
    put(ws5, row, [None] * 4, fmt=[None, "#,##0", "#,##0", "#,##0"])
    row += 1

for s in wb.worksheets:
    s.sheet_properties.tabColor = "1F3864"
OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
print("saved:", OUT)
