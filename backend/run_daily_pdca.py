#!/usr/bin/env python3
"""日次 PDCA ランナー — Check（計測）→ 分析レポート → Act（テーマキュー調整）を自動実行。

設計方針:
    動画投稿のオートパイロットと同様、テーマキューや analytics.db を持つのは
    「ライブのバックエンドサーバ（localhost:8000）」である。channel_manager は
    _raw をメモリにキャッシュしており、外部から data/channels/*.json を直接書くと
    サーバ側の保存と競合してテーマが消える（過去に dup-flood を起こした経路）。

    そのため本ランナーは **すべての更新系をライブサーバの HTTP API 経由で行う**。
    こうすればサーバのインメモリ状態が常に正となり、競合は起きない。

実行内容（チャンネルごと、analytics.enabled=true のみ）:
    1. Check  : POST /api/analytics/sync/{id}
                 → YouTube から再生数・いいね・維持率を取得して analytics.db へ。
                   付随して scenario評価 / AB答え合わせ / 改善キュー / シリーズ検出 /
                   コメント需要抽出（サーバ側 PDCA チェーン）も走る。
    2. 分析   : pdca-report / videos / optimal-posting-time / series を取得して
                 - 動画ごとの再生数推移
                 - テーマ（ジャンル）別の成績
                 - 投稿時間帯 × 再生数の相関
                 - テーマ重複チェック
                をまとめる。
    3. Act    : 「高再生のバズ動画」から生成された続編候補（series_suggestions）を
                 一定の絶対再生数を超えたものだけ自動承認し、テーマキュー先頭へ投入。
                 低再生ジャンルは抑制候補としてレポートに記載（破壊的操作はしない）。
    4. 保存   : data/reports/YYYY-MM-DD/ に JSON + Markdown を出力。

毎日 23:00(JST) に launchd（com.youtube-factory.pdca）から起動される想定。
手動実行: `python3 backend/run_daily_pdca.py [channel_id ...]`
"""

from __future__ import annotations

import json
import os
import sqlite3
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

BACKEND_DIR = Path(__file__).resolve().parent
REPO_ROOT = BACKEND_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))

# --- .env 読み込み（JWT/APP_PASSWORD/暗号鍵などに必要）---------------------
env_file = BACKEND_DIR / ".env"
if env_file.exists():
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip())

JST = timezone(timedelta(hours=9))
BASE_URL = os.environ.get("PDCA_BASE_URL", "http://localhost:8000").rstrip("/")
CHANNELS_DIR = REPO_ROOT / "data" / "channels"
REPORTS_DIR = REPO_ROOT / "data" / "reports"
ANALYTICS_DB = REPO_ROOT / "data" / "analytics" / "analytics.db"

# Act（自動承認）のガード — 初期の超低再生ノイズで続編が量産されるのを防ぐ
MIN_VIRAL_VIEWS = 30      # この絶対再生数を超えたバズ動画の続編のみ自動投入
MIN_VIRAL_RATIO = 1.5     # チャンネル平均比
MAX_APPROVALS_PER_RUN = 2  # 1チャンネル/日あたりの自動投入上限
DEDUP_SIM_THRESHOLD = 0.62  # テーマ重複とみなす類似度

# 分析対象の期間
SYNC_DAYS = 30

# ジャンル分類は生成側（genre_blacklist）と共有する。レポートで「平均0再生」と出た
# ジャンル名をそのままチャンネル JSON の genre_blacklist に書けるよう、分類器は 1 つに保つ。
from pipeline.auto_scenario.genre import GENRE_KEYWORDS, classify_genre as _classify_genre  # noqa: E402


# =====================================================================
# HTTP ヘルパ
# =====================================================================

def _req(method: str, path: str, token: Optional[str] = None,
         body: Optional[dict] = None, timeout: float = 180.0) -> Any:
    url = path if path.startswith("http") else f"{BASE_URL}{path}"
    data = None
    headers = {"Accept": "application/json"}
    if body is not None:
        data = json.dumps(body).encode("utf-8")
        headers["Content-Type"] = "application/json"
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read().decode("utf-8")
    return json.loads(raw) if raw else {}


def _login() -> str:
    pw = os.environ.get("APP_PASSWORD", "")
    if not pw:
        raise SystemExit("APP_PASSWORD が .env にありません — ログインできません")
    res = _req("POST", "/api/auth/login", body={"password": pw}, timeout=15)
    token = res.get("token")
    if not token:
        raise SystemExit(f"ログイン失敗: {res}")
    return token


def _safe(method: str, path: str, token: str, body: Optional[dict] = None,
          timeout: float = 180.0) -> Dict[str, Any]:
    """例外を握って {ok, data|error} に正規化する。"""
    try:
        data = _req(method, path, token=token, body=body, timeout=timeout)
        return {"ok": True, "data": data}
    except urllib.error.HTTPError as e:
        detail = ""
        try:
            detail = e.read().decode("utf-8")[:300]
        except Exception:
            pass
        return {"ok": False, "error": f"HTTP {e.code} {path} {detail}"}
    except Exception as e:
        return {"ok": False, "error": f"{type(e).__name__}: {e}"}


# =====================================================================
# 分析ヘルパ
# =====================================================================

def _enabled_channels(explicit: List[str]) -> List[str]:
    ids: List[str] = []
    for f in sorted(CHANNELS_DIR.glob("*.json")):
        try:
            d = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        cid = d.get("id") or f.stem
        if explicit and cid not in explicit:
            continue
        enabled = (((d.get("video_format") or {}).get("analytics") or {}).get("enabled"))
        if enabled or explicit:  # 明示指定があれば enabled に関わらず対象にする
            ids.append(cid)
    return ids


def _genre_breakdown(channel_id: str, videos: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    buckets: Dict[str, Dict[str, Any]] = {}
    for v in videos:
        g = _classify_genre(channel_id, v.get("title") or "")
        b = buckets.setdefault(g, {"genre": g, "count": 0, "total_views": 0,
                                    "total_likes": 0, "titles": []})
        b["count"] += 1
        b["total_views"] += int(v.get("views") or 0)
        b["total_likes"] += int(v.get("likes") or 0)
        if len(b["titles"]) < 3:
            b["titles"].append(v.get("title") or "")
    out = []
    for b in buckets.values():
        b["avg_views"] = round(b["total_views"] / b["count"], 2) if b["count"] else 0.0
        out.append(b)
    out.sort(key=lambda x: x["avg_views"], reverse=True)
    return out


def _view_trends(channel_id: str, max_videos: int = 40,
                 live_views: Optional[Dict[str, int]] = None) -> List[Dict[str, Any]]:
    """analytics.db を read-only で読み、動画ごとの (date, views) 推移を返す。"""
    live_views = live_views or {}
    if not ANALYTICS_DB.exists():
        return []
    try:
        uri = f"file:{ANALYTICS_DB}?mode=ro"
        conn = sqlite3.connect(uri, uri=True, timeout=10)
        conn.row_factory = sqlite3.Row
    except Exception:
        return []
    try:
        rows = conn.execute(
            "SELECT video_id, title, date, views, likes FROM video_metrics "
            "WHERE channel_id = ? ORDER BY video_id, date",
            (channel_id,),
        ).fetchall()
    except Exception:
        return []
    finally:
        conn.close()
    by_video: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        vid = r["video_id"]
        e = by_video.setdefault(vid, {"video_id": vid, "title": r["title"], "series": []})
        e["title"] = r["title"] or e["title"]
        e["series"].append({"date": r["date"], "views": int(r["views"] or 0),
                            "likes": int(r["likes"] or 0)})
    out = []
    for e in by_video.values():
        s = e["series"]
        snap_latest = s[-1]["views"] if s else 0
        e["latest_views"] = max(snap_latest, int(live_views.get(e["video_id"], 0)))
        e["first_views"] = s[0]["views"] if s else 0
        e["delta"] = e["latest_views"] - e["first_views"]
        e["snapshots"] = len(s)
        out.append(e)
    out.sort(key=lambda x: x["latest_views"], reverse=True)
    return out[:max_videos]


def _dup_check(channel_id: str, videos: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """published 済みタイトル同士の意味的に近いペアを抽出（テーマ重複の検知）。"""
    try:
        from pipeline.auto_scenario import theme_dedup as td
    except Exception:
        return []
    titles = [v.get("title") or "" for v in videos if v.get("title")]
    pairs: List[Dict[str, Any]] = []
    for i in range(len(titles)):
        for j in range(i + 1, len(titles)):
            try:
                sim = td.similarity(titles[i], titles[j])
            except Exception:
                continue
            if sim >= DEDUP_SIM_THRESHOLD:
                pairs.append({"a": titles[i], "b": titles[j], "similarity": round(sim, 3)})
    pairs.sort(key=lambda x: x["similarity"], reverse=True)
    return pairs[:15]


# =====================================================================
# Act — バズ続編の自動投入
# =====================================================================

def _act_promote(channel_id: str, token: str, series: Dict[str, Any]) -> Dict[str, Any]:
    grouped = sorted(
        series.get("grouped") or [],
        key=lambda g: float(g.get("viral_ratio") or 0.0),
        reverse=True,
    )
    approved: List[Dict[str, Any]] = []
    skipped_low: List[Dict[str, Any]] = []
    for g in grouped:
        if len(approved) >= MAX_APPROVALS_PER_RUN:
            break
        views = int(g.get("original_views") or 0)
        ratio = float(g.get("viral_ratio") or 0.0)
        if views < MIN_VIRAL_VIEWS or ratio < MIN_VIRAL_RATIO:
            skipped_low.append({"title": g.get("original_title"), "views": views, "ratio": ratio})
            continue
        for s in (g.get("suggestions") or []):
            if len(approved) >= MAX_APPROVALS_PER_RUN:
                break
            if (s.get("status") or "pending") != "pending":
                continue
            sid = s.get("suggestion_id") or s.get("id")
            if not sid:
                continue
            res = _safe("POST", f"/api/series/{channel_id}/approve/{sid}", token, timeout=30)
            if res.get("ok") and (res["data"] or {}).get("ok", True) is not False:
                approved.append({
                    "from": g.get("original_title"),
                    "from_views": views,
                    "viral_ratio": ratio,
                    "series_type": s.get("series_type"),
                    "title": s.get("suggested_title"),
                })
            else:
                approved.append({"error": res.get("error") or (res.get("data")), "sid": sid})
    return {"approved": approved, "skipped_low_view": skipped_low[:10]}


# =====================================================================
# レポート整形
# =====================================================================

def _fmt_pct(x: Optional[float]) -> str:
    return f"{x*100:.1f}%" if isinstance(x, (int, float)) else "—"


def _channel_markdown(rep: Dict[str, Any]) -> str:
    cid = rep["channel_id"]
    L: List[str] = []
    L.append(f"## {rep.get('channel_name', cid)} (`{cid}`)\n")

    sync = rep.get("sync") or {}
    if not sync.get("ok"):
        L.append(f"> ⚠️ Check(sync) 失敗: {sync.get('error')}\n")

    # サマリ（pdca-report より）
    pr = (rep.get("pdca_report") or {}).get("data") or {}
    cs = pr.get("channel_stats") or {}
    sh = pr.get("shorts") or {}
    rec = pr.get("recommendation") or {}
    L.append("### サマリ")
    L.append(f"- 登録者: **{cs.get('subscriber_count', '—')}** / 総再生: {cs.get('view_count', '—')} / 本数: {cs.get('video_count', '—')}")
    L.append(f"- 直近{SYNC_DAYS}日 ショート: {sh.get('count', 0)}本 / 平均再生 {sh.get('avg_views', 0)} / 中央値 {sh.get('median_views', 0)} / 平均いいね率 {_fmt_pct(sh.get('avg_like_rate'))}")
    if rec.get("headline"):
        L.append(f"- 判定: **{rec.get('headline')}**")
    L.append("")

    # 再生数 上位/推移
    trends = rep.get("view_trends") or []
    L.append("### 動画ごとの再生数（推移スナップショット）")
    if trends:
        L.append("| 再生 | Δ | 計測日数 | タイトル |")
        L.append("|---:|---:|---:|---|")
        for t in trends[:15]:
            title = (t.get("title") or "")[:42]
            L.append(f"| {t.get('latest_views',0)} | {t.get('delta',0):+d} | {t.get('snapshots',0)} | {title} |")
    else:
        L.append("_計測データなし（sync 後に蓄積されます）_")
    L.append("")

    # ジャンル別
    gb = rep.get("genre_breakdown") or []
    L.append("### テーマ（ジャンル）別の成績")
    if gb:
        L.append("| ジャンル | 本数 | 平均再生 | 合計再生 | 合計いいね |")
        L.append("|---|---:|---:|---:|---:|")
        for b in gb:
            L.append(f"| {b['genre']} | {b['count']} | {b['avg_views']} | {b['total_views']} | {b['total_likes']} |")
        top = gb[0]["genre"] if gb else "—"
        low = gb[-1]["genre"] if len(gb) > 1 else "—"
        L.append("")
        L.append(f"- 🔼 伸びてる系統: **{top}** → 続編・派生を優先")
        L.append(f"- 🔽 抑制候補: **{low}**（平均再生が最下位）")
    else:
        L.append("_分類できる動画がまだありません_")
    L.append("")

    # 投稿時間帯
    po = ((rep.get("posting") or {}).get("data") or {})
    porec = (po.get("recommendation") or {}).get("recommended") or {}
    L.append("### 投稿時間帯 × 再生数の相関")
    if porec:
        dow_names = ["日", "月", "火", "水", "木", "金", "土"]
        d = porec.get("day_of_week")
        dn = dow_names[d] if isinstance(d, int) and 0 <= d <= 6 else "—"
        L.append(f"- 実績ベスト: **{dn}曜 {porec.get('hour','—')}:00**（平均 {porec.get('avg_views','—')}再生 / サンプル {porec.get('sample_size','—')}本 / 平均比 {porec.get('boost_percent','—')}%）")
        if porec.get("sample_size", 0) < 3:
            L.append("- ⚠️ サンプル数が少なく統計的信頼性は低い（参考値）。固定スロット(12:00/19:00)は維持中。")
    else:
        L.append("_データ不足_")
    L.append("")

    # 重複チェック
    dups = rep.get("dup_check") or []
    L.append("### テーマ重複チェック")
    if dups:
        L.append(f"- 類似ペア {len(dups)} 件（閾値 {DEDUP_SIM_THRESHOLD}）:")
        for p in dups[:8]:
            L.append(f"  - `{p['similarity']}` {p['a'][:30]} ↔ {p['b'][:30]}")
    else:
        L.append("- 重複疑いなし ✅")
    L.append("")

    # 成功パターン分析
    sp = rep.get("success_patterns") or {}
    L.append("### 成功パターン分析 → シナリオ生成フィードバック")
    if sp.get("ok"):
        ss = sp.get("sample_size") or {}
        L.append(f"- 成功動画: **{ss.get('success', '—')}**本 / 全体: {ss.get('total', '—')}本")
        if sp.get("has_gpt_insights"):
            L.append("- Claude分析: ✅ 完了（actionable_recommendations が scenario_feedback 経由で次回シナリオ生成に自動注入）")
        else:
            L.append(f"- Claude分析: スキップ（{sp.get('gpt_skipped_reason') or '理由不明'}）")
    else:
        L.append(f"- ⚠️ 分析失敗: {sp.get('error', '不明')}")
    L.append("")

    # 維持率分析
    ri = rep.get("retention_insights") or {}
    L.append("### 視聴維持率分析 → シナリオ生成フィードバック")
    if ri.get("ok"):
        if ri.get("has_gpt_insights"):
            L.append("- Claude分析: ✅ 完了（retention_tips が scenario_feedback 経由で次回シナリオ生成に自動注入）")
        else:
            L.append(f"- Claude分析: スキップ（{ri.get('gpt_skipped_reason') or '理由不明'}）")
    else:
        L.append(f"- ⚠️ 分析失敗: {ri.get('error', '不明')}")
    L.append("")

    # Act
    act = rep.get("act") or {}
    L.append("### Act（高再生テーマの続編をキュー投入）")
    appr = act.get("approved") or []
    if appr:
        for a in appr:
            if a.get("error"):
                L.append(f"- ⚠️ 承認失敗: {a['error']}")
            else:
                L.append(f"- ✅ 投入: 「{a.get('title')}」（{a.get('series_type')} / 元動画 {a.get('from_views')}再生・平均比{a.get('viral_ratio'):.1f}倍）")
    else:
        L.append(f"- 投入なし（絶対再生 {MIN_VIRAL_VIEWS} 以上のバズ続編が未検出）")
    L.append("")
    return "\n".join(L)


# =====================================================================
# xlsx 履歴（チャンネルごとに1シート、1日1行を追記）
# =====================================================================

HISTORY_XLSX = REPORTS_DIR / "pdca_history.xlsx"

# 列の並び（ヘッダー）。日付をキーにして冪等に更新する。
XLSX_COLUMNS = [
    "日付", "登録者数", "総再生数", "動画本数", "直近30日ショート本数",
    "平均再生数", "中央値再生数", "平均いいね率", "伸びてるジャンル",
    "抑制候補ジャンル", "テーマ重複ペア数", "成功パターン分析(OK/NG)",
    "維持率分析(OK/NG)", "バズ続編投入数",
]


def _rep_to_row(rep: Dict[str, Any], date_str: str) -> List[Any]:
    """rep 辞書から xlsx の1行分（XLSX_COLUMNS の順）を作る。"""
    pr = (rep.get("pdca_report") or {}).get("data") or {}
    cs = pr.get("channel_stats") or {}
    sh = pr.get("shorts") or {}
    gb = rep.get("genre_breakdown") or []
    dups = rep.get("dup_check") or []
    sp = rep.get("success_patterns") or {}
    ri = rep.get("retention_insights") or {}
    act = rep.get("act") or {}

    top_genre = gb[0]["genre"] if gb else ""
    low_genre = gb[-1]["genre"] if len(gb) > 1 else ""
    like_rate = sh.get("avg_like_rate")  # 生の比率（0.015 = 1.5%）。セル側で % 表示。
    approved = [a for a in (act.get("approved") or []) if not a.get("error")]

    return [
        date_str,
        cs.get("subscriber_count"),
        cs.get("view_count"),
        cs.get("video_count"),
        sh.get("count", 0),
        sh.get("avg_views", 0),
        sh.get("median_views", 0),
        like_rate if isinstance(like_rate, (int, float)) else None,
        top_genre,
        low_genre,
        len(dups),
        "OK" if sp.get("ok") else "NG",
        "OK" if ri.get("ok") else "NG",
        len(approved),
    ]


def _write_xlsx_history(reports: List[Dict[str, Any]], date_str: str) -> None:
    """pdca_history.xlsx にチャンネルごと1シートで当日の行を追記/更新する。

    - ファイルが無ければ新規作成、あれば読み込んで追記。
    - 同じ日付の行が既にあれば上書き（同日再実行で重複させない）。
    - ヘッダーは太字＋色付け、オートフィルタ＋ウィンドウ固定。
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception as e:  # openpyxl 未導入でも PDCA 本体は止めない
        print(f"  [xlsx] スキップ: openpyxl 読み込み失敗 ({e})")
        return

    HISTORY_XLSX.parent.mkdir(parents=True, exist_ok=True)

    if HISTORY_XLSX.exists():
        wb = load_workbook(HISTORY_XLSX)
    else:
        wb = Workbook()
        # 既定の空シートは後で使うチャンネルシートに置き換えるため削除予約
        wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    like_col_idx = XLSX_COLUMNS.index("平均いいね率") + 1  # 1-based

    for rep in reports:
        cid = rep.get("channel_id") or "unknown"
        sheet_name = cid[:31]  # Excel のシート名は31文字上限

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(XLSX_COLUMNS)

        # ヘッダーが欠けている場合の保険（既存ファイルの空シート等）
        if ws.max_row == 0 or (ws.cell(row=1, column=1).value != XLSX_COLUMNS[0]):
            ws.insert_rows(1)
            for c, name in enumerate(XLSX_COLUMNS, start=1):
                ws.cell(row=1, column=c, value=name)

        row = _rep_to_row(rep, date_str)

        # 同日行を探して上書き、無ければ追記
        target_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == date_str:
                target_row = r
                break
        if target_row is None:
            target_row = ws.max_row + 1
        for c, val in enumerate(row, start=1):
            ws.cell(row=target_row, column=c, value=val)

        # 見た目の整形
        for c, name in enumerate(XLSX_COLUMNS, start=1):
            hc = ws.cell(row=1, column=c)
            hc.font = header_font
            hc.fill = header_fill
            hc.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(c)].width = max(12, len(name) + 2)
        # いいね率列を % 表示に
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=like_col_idx).number_format = "0.00%"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(XLSX_COLUMNS))}{ws.max_row}"

    wb.save(HISTORY_XLSX)
    print(f"  [xlsx] 履歴更新: {HISTORY_XLSX} （{len(reports)}ch / {date_str}）")


# =====================================================================
# メイン
# =====================================================================

def run_channel(channel_id: str, channel_name: str, token: str,
                do_sync: bool = True) -> Dict[str, Any]:
    print(f"\n=== PDCA: {channel_id} ===")
    rep: Dict[str, Any] = {"channel_id": channel_id, "channel_name": channel_name}

    # 1. Check — sync（重い。最大数十秒〜数分）
    if do_sync:
        print("  [Check] sync...")
        sync_body = {
            "days": SYNC_DAYS,
            "max_videos": 50,
            "fetch_retention_for": 5,
            "sync_comments_for": 5,
            "run_scenario_evaluation": True,
            "run_ab_reconciliation": True,
            "run_improvement_detection": True,
            "run_series_detection": True,
            "run_comment_demand": True,
        }
        sync = _safe("POST", f"/api/analytics/sync/{channel_id}", token, body=sync_body, timeout=600)
        rep["sync"] = sync
        if sync.get("ok"):
            v = ((sync["data"] or {}).get("videos") or {})
            print(f"    sync ok — videos: {len(v.get('items') or [])}")
        else:
            print(f"    sync FAILED: {sync.get('error')}")
    else:
        print("  [Check] sync スキップ (--no-sync)")
        rep["sync"] = {"ok": True, "skipped": True}

    # 2. 分析データ取得
    print("  [Analyze] reports...")
    rep["pdca_report"] = _safe("GET", f"/api/analytics/pdca-report?channel_id={channel_id}&days={SYNC_DAYS}", token, timeout=120)
    rep["posting"] = _safe("GET", f"/api/optimal-posting-time/{channel_id}?days={SYNC_DAYS}&recompute=true", token, timeout=120)
    series = _safe("GET", f"/api/series/{channel_id}", token, timeout=120)
    rep["series"] = series

    videos_res = _safe("GET", f"/api/analytics/videos/{channel_id}?limit=200", token, timeout=60)
    videos = ((videos_res.get("data") or {}).get("items") or []) if videos_res.get("ok") else []

    # YouTube Analytics API は新着動画の views を遅延反映する（0 になりがち）。
    # pdca-report は Data API の statistics で「現在の実再生数」を持つので、これを
    # マージしてジャンル分析・推移を即座に意味あるものにする。
    live_views: Dict[str, int] = {}
    live_likes: Dict[str, int] = {}
    pr = (rep["pdca_report"].get("data") or {}) if rep["pdca_report"].get("ok") else {}
    for bucket in ((pr.get("shorts") or {}).get("top_videos") or []) + ((pr.get("main") or {}).get("top_videos") or []):
        vid = bucket.get("video_id")
        if vid:
            live_views[vid] = int(bucket.get("views") or 0)
            live_likes[vid] = int(bucket.get("likes") or 0)
    for v in videos:
        vid = v.get("video_id")
        if vid in live_views:
            v["views"] = max(int(v.get("views") or 0), live_views[vid])
            v["likes"] = max(int(v.get("likes") or 0), live_likes.get(vid, 0))

    rep["genre_breakdown"] = _genre_breakdown(channel_id, videos)
    rep["view_trends"] = _view_trends(channel_id, live_views=live_views)
    rep["dup_check"] = _dup_check(channel_id, videos)

    # 2b. 成功パターン分析 + 視聴維持率分析 — シナリオ生成へのフィードバックを更新
    print("  [Analyze] success patterns & retention...")
    try:
        from pipeline.analytics import success_analyzer
        sp = success_analyzer.analyze_channel(channel_id, use_gpt=True)
        rep["success_patterns"] = {
            "ok": True,
            "sample_size": sp.get("sample_size"),
            "has_gpt_insights": sp.get("gpt_insights") is not None,
            "gpt_skipped_reason": sp.get("gpt_skipped_reason"),
        }
        recs = (sp.get("gpt_insights") or {}).get("actionable_recommendations") or []
        print(f"    success patterns ok — {sp.get('sample_size', {}).get('success', 0)} success videos, {len(recs)} recommendations")
    except Exception as e:
        rep["success_patterns"] = {"ok": False, "error": str(e)}
        print(f"    success patterns FAILED: {e}")

    try:
        from pipeline.analytics import retention_analyzer
        ri = retention_analyzer.analyze_channel(channel_id, use_gpt=True, max_videos=10)
        rep["retention_insights"] = {
            "ok": True,
            "has_gpt_insights": ri.get("gpt_insights") is not None,
            "gpt_skipped_reason": ri.get("gpt_skipped_reason"),
        }
        tips = (ri.get("gpt_insights") or {}).get("retention_tips") or []
        print(f"    retention insights ok — {len(tips)} tips")
    except Exception as e:
        rep["retention_insights"] = {"ok": False, "error": str(e)}
        print(f"    retention insights FAILED: {e}")

    # 3. Act — バズ続編の自動投入
    print("  [Act] promote viral sequels...")
    if series.get("ok"):
        rep["act"] = _act_promote(channel_id, token, series["data"] or {})
        print(f"    approved: {len(rep['act'].get('approved') or [])}")
    else:
        rep["act"] = {"approved": [], "error": series.get("error")}

    return rep


def main(argv: List[str]) -> int:
    explicit = [a for a in argv[1:] if not a.startswith("-")]
    do_sync = "--no-sync" not in argv
    now = datetime.now(JST)
    date_str = now.strftime("%Y-%m-%d")
    out_dir = REPORTS_DIR / date_str
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        token = _login()
    except SystemExit as e:
        print(str(e))
        return 2

    channel_ids = _enabled_channels(explicit)
    if not channel_ids:
        print("対象チャンネルなし（analytics.enabled=true のチャンネルが無い）")
        return 0
    print(f"対象チャンネル: {channel_ids}")

    # channel 名取得（任意）
    names: Dict[str, str] = {}
    for cid in channel_ids:
        try:
            d = json.loads((CHANNELS_DIR / f"{cid}.json").read_text(encoding="utf-8"))
            names[cid] = d.get("name") or cid
        except Exception:
            names[cid] = cid

    reports: List[Dict[str, Any]] = []
    md_parts: List[str] = [
        f"# 日次 PDCA レポート — {date_str}",
        f"_生成: {now.strftime('%Y-%m-%d %H:%M %Z')} / base={BASE_URL}_\n",
    ]
    for cid in channel_ids:
        try:
            rep = run_channel(cid, names[cid], token, do_sync=do_sync)
        except Exception as e:
            import traceback
            traceback.print_exc()
            rep = {"channel_id": cid, "channel_name": names[cid], "fatal": str(e)}
        reports.append(rep)
        # per-channel JSON
        (out_dir / f"{cid}.json").write_text(
            json.dumps(rep, ensure_ascii=False, indent=2), encoding="utf-8")
        md_parts.append(_channel_markdown(rep))

    combined_md = "\n".join(md_parts)
    (out_dir / "report.md").write_text(combined_md, encoding="utf-8")
    (REPORTS_DIR / "latest.md").write_text(combined_md, encoding="utf-8")

    # xlsx 履歴に当日分を追記（チャンネルごと1シート）
    try:
        _write_xlsx_history(reports, date_str)
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"⚠️ xlsx 履歴の書き込みに失敗: {e}")
    (out_dir / "summary.json").write_text(
        json.dumps({"date": date_str, "generated_at": now.isoformat(),
                    "channels": channel_ids}, ensure_ascii=False, indent=2),
        encoding="utf-8")

    print(f"\n✅ レポート出力: {out_dir}/report.md")
    print(f"   最新版: {REPORTS_DIR / 'latest.md'}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
